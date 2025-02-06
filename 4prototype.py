# -*- coding: utf-8 -*-

import sys
import pandas as pd
import matplotlib.pyplot as plt
from PyQt6.QtWidgets import QApplication, QInputDialog, QAbstractItemView, QTableView, QMainWindow, QVBoxLayout, QComboBox, QAbstractItemView, QFileDialog, QMessageBox
from PyQt6.QtCore import Qt, QAbstractTableModel, QModelIndex, pyqtSignal
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.figure import Figure
from PyQt6.QtCore import QAbstractTableModel, QModelIndex, Qt
from PyQt6.uic import loadUi
from scipy.io import loadmat
from scipy.signal import find_peaks
import numpy as np
from scipy.ndimage import gaussian_filter1d
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

# Класс для модели данных Pandas
class PandasModel(QAbstractTableModel):
    def __init__(self, data):
        super().__init__()
        self._data = data
        self._headers = ["Experiment No.", "Start Experiment (s)", "Hand Lifting (s)", "QA Opening", "Object Lifting (s)", "QA max", "Object placed"]

    def rowCount(self, parent=QModelIndex()):
        return self._data.shape[0]

    def columnCount(self, parent=QModelIndex()):
        return self._data.shape[1]

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if index.isValid() and role == Qt.ItemDataRole.DisplayRole:
            return str(self._data.iloc[index.row(), index.column()])
        return None
    
    def headerData(self, section, orientation, role):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                return self._headers[section] 
        return None

    def setHorizontalHeaderLabels(self, headers):
        self._headers = headers

# Главное окно приложения
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        #loadUi("pril.ui", self)  # Загрузка пользовательского интерфейса

        ui_file = self.resource_path("pril.ui")
        loadUi(ui_file, self)

        self.file_loaded = False # проверка на то, что файл в программе имеется

        # Значения по умолчанию
        self.default_values = {
            'tl': 410,  # расстояние между пиками
            'hand_lifting': 0.005,  # скорость руки
            'finger_opening': 1.2,  # открытие пальца
            'object_lifting': 0.01,  # подъем объекта
            'object_lowering': 0.005,  # опускание предмета
            'max_FG': 0.07,  # максимум Frame-Glasses
            'min_FG': 0.051  # минимум Frame-Glasses
        }

        # Устанавливаем значения на текущие
        self.tl = self.default_values['tl']
        self.hand_lifting = self.default_values['hand_lifting']
        self.finger_opening = self.default_values['finger_opening']
        self.object_lifting = self.default_values['object_lifting']
        self.object_lowering = self.default_values['object_lowering']
        self.max_FG = self.default_values['max_FG']
        self.min_FG = self.default_values['min_FG']

        #создаем модель для графиков
        self.figure = Figure()
        self.canvas = FigureCanvas(self.figure)

        #добавляем Canvas в GraphicsView
        layout = QVBoxLayout(self.graphicsView)
        layout.addWidget(self.canvas)

        #навигационная панель
        self.toolbar = NavigationToolbar(self.canvas, self)
        layout.addWidget(self.toolbar)
        self.graphicsView.setLayout(layout)

        # Подключение кнопок к соответствующим методам
        self.pushButton.clicked.connect(self.import_data)
        self.pushButton_3.clicked.connect(self.export_data_to_excel)
        self.pushButton_6.clicked.connect(self.on_plot_button_clicked)
        self.pushButton_2.clicked.connect(self.open_settings) # Настройки
        self.pushButton_7.clicked.connect(self.on_plot_button_clicked) # Кнопка перестройки графика

        # Поиск и инициализация TableViТew
        self.tableView = self.findChild(QTableView, 'tableView')

        #Выделение всей строки при выборе
        self.tableView.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)

    def resource_path(self, relative_path):
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, relative_path)
        return os.path.join(os.path.abspath("."), relative_path)

    def open_settings(self):
        self.settings_window = QMainWindow()
        #loadUi('settings.ui', self.settings_window)  

        ui_file = self.resource_path("settings.ui")
        loadUi(ui_file, self.settings_window)

        # Установка текущих значений в настройки
        self.settings_window.lineEdit.setText(str(self.tl))  # Расстояние между пиками
        self.settings_window.lineEdit_2.setText(str(self.hand_lifting))  # Порог скорости руки
        self.settings_window.lineEdit_3.setText(str(self.finger_opening))  # Порог открытия пальца
        self.settings_window.lineEdit_5.setText(str(self.object_lifting))  # Порог подъема объекта
        self.settings_window.lineEdit_4.setText(str(self.object_lowering))  # Порог опускания объекта
        self.settings_window.lineEdit_6.setText(str(self.max_FG))  # Максимум FG
        self.settings_window.lineEdit_7.setText(str(self.min_FG))  # Минимум FG

        self.settings_window.show() 
        self.settings_window.pushButton_4.clicked.connect(self.save_settings)
        self.settings_window.pushButton_5.clicked.connect(self.return_defoult_data)

    def save_settings(self):
        # Сохранение настроек при закрытии окна
        try:
            self.tl = float(self.settings_window.lineEdit.text().replace(',', '.'))  # расстояние между пиками
            self.hand_lifting = float(self.settings_window.lineEdit_2.text().replace(',', '.'))  # Порог скорости руки
            self.finger_opening = float(self.settings_window.lineEdit_3.text().replace(',', '.'))  # Порог открытия пальца
            self.object_lifting = float(self.settings_window.lineEdit_5.text().replace(',', '.'))  # Порог подъема объекта
            self.object_lowering = float(self.settings_window.lineEdit_4.text().replace(',', '.'))  # Порог опускания объекта
            self.max_FG = float(self.settings_window.lineEdit_6.text().replace(',', '.'))  # Максимум FG
            self.min_FG = float(self.settings_window.lineEdit_7.text().replace(',', '.'))  # Минимум FG

            self.update_experiment_table()
        except ValueError:
            QMessageBox.warning(self, "Ошибка", "Введены некорректные значения в настройках")

    # Вернуть данные по умолчанию
    def return_defoult_data(self): 

        # Возвращаем значения по умолчанию
        self.settings_window.lineEdit.setText(str(self.default_values['tl']))  # Расстояние между пиками
        self.settings_window.lineEdit_2.setText(str(self.default_values['hand_lifting']))  # Порог скорости руки
        self.settings_window.lineEdit_3.setText(str(self.default_values['finger_opening']))  # Порог открытия пальца
        self.settings_window.lineEdit_5.setText(str(self.default_values['object_lifting']))  # Порог подъема объекта
        self.settings_window.lineEdit_4.setText(str(self.default_values['object_lowering']))  # Порог опускания объекта
        self.settings_window.lineEdit_6.setText(str(self.default_values['max_FG']))  # Максимум FG
        self.settings_window.lineEdit_7.setText(str(self.default_values['min_FG']))  # Минимум FG

        # Обновляем текущие значения
        self.tl = self.default_values['tl']
        self.hand_lifting = self.default_values['hand_lifting']
        self.finger_opening = self.default_values['finger_opening']
        self.object_lifting = self.default_values['object_lifting']
        self.object_lowering = self.default_values['object_lowering']
        self.max_FG = self.default_values['max_FG']
        self.min_FG = self.default_values['min_FG']
        

    def import_data(self):
        self.file_path, _ = QFileDialog.getOpenFileName(self, "MAT files", "*.mat")
        
        if not self.file_path:
            QMessageBox.warning(self, "Предупреждение", "Файл не выбран.")
            return -1
        try:
            self.data = loadmat(self.file_path)
            self.file_loaded = True

            if not self.data:
                QMessageBox.warning(self, "Ошибка", "Некорректный файл (проверьте содержимой файла)")
        except FileNotFoundError:
            QMessageBox.warning(self, "Ошибка", "Файл не найден. Пожалуйста, проверьту путь к файлу.")
            return -1
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Произошла ошибка при загрузке файла: {e}")
            return - 1
    
        
        # Сохраняем путь к файлу
        self.current_file_path = self.file_path

        self.calculate_parameters()
        self.moment_1_start_experiment()
        self.moment_2_hand_lifting()
        self.moment_3_finger_opening()
        self.moment_4_object_lifting()
        self.moment_5_max_ga()
        self.moment_6_object_lowering()
        self.create_experiment_array()

        # Создание таблицы с данными экспериментов
        self.experiment_table = None  # Инициализация таблицы
        self.create_experiment_table()  # Создаем таблицу после загрузки данных
        
    def update_experiment_table(self):
        """Обновляет данные в отчете, используя путь к загруженному файлу."""
        try:
            # Загружаем данные из сохраненного пути
            self.data = loadmat(self.current_file_path)
            self.calculate_parameters()
            self.moment_1_start_experiment()
            self.moment_2_hand_lifting()
            self.moment_3_finger_opening()
            self.moment_4_object_lifting()
            self.moment_5_max_ga()
            self.moment_6_object_lowering()
            self.create_experiment_array()
            # Создание таблицы с данными экспериментов
            self.experiment_table = None  # Инициализация таблицы
            self.create_experiment_table()  # Создаем таблицу после загрузки данных
            if not hasattr(self, 'experiment_data') or not self.experiment_data:
                self.figure.clear() 
                self.canvas.draw()
                QMessageBox.warning(self, "Ошибка", "Отсутствуют данные, соответствующие указанным настройкам (измените настройки параметров на настройки по умолчанию)")

            else:
                QMessageBox.information(self, "Уведомление", "Данные успешно обновлены!", QMessageBox.StandardButton.Ok)
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Произошла ошибка при обновлении данных: {e}")

    def create_experiment_table(self):
        """Создает таблицу экспериментов в интерфейсе."""
        self.generate_experiment_data()  # Генерируем данные для таблицы

        # Создаем Dataframe
        df = pd.DataFrame(self.experiment_data)
        
        # Натсраиваем модель для табличного воспроизведения
        self.model = PandasModel(df)
        self.tableView.setModel(self.model)

        # заголовки таблицы
        self.model.setHorizontalHeaderLabels(["№", "Start Experiment (s)", "Hand Lifting (s)", "GA Opening", "Object Lifting (s)", "GA max", "Object placed"])

        #Установка стиля для заголовка
        header = self.tableView.horizontalHeader()
        header.setStyleSheet("QHeaderView::section { background-color: lightgrey; border-bottom: 1px solid black; }")

    def on_plot_button_clicked(self):
        if self.file_loaded == False:
            QMessageBox.warning(self, "Ошибка", "Отсутствуют данные для работы (загрузите файл)")
        else:
            if not hasattr(self, 'experiment_data') or not self.experiment_data:
                QMessageBox.warning(self, "Ошибка", "Отсутствуют данные соответствующие указанным настройкам (измените настройки параметров на настройки по умолчанию)")
                return
            else:
                selected_row = self.get_selected_row()
                if selected_row != -1:
                    self.plot_results_experiment(selected_row)
                else:
                    QMessageBox.warning(self, "Ошибка", "Выберите эксперимент для визуализации, нажав на соответствующую строку в таблице")

    def get_selected_row(self):
        # Получаем модель выбора
        selection_model = self.tableView.selectionModel()
        # Получаем индекс выбранного ряда
        selected_indexes = selection_model.selectedRows()
        if selected_indexes:
            return selected_indexes[0].row()  # Возвращаем индекс первого выбранного ряда
        return -1  # Если ничего не выбрано, возвращаем -1
        
    
    def generate_experiment_data(self):
        """Формирует данные для каждого эксперимента."""
        self.experiment_data = []

        for idx, (start_point, end_point) in enumerate(self.experiment_array):
            # выборка точки времени начала эксперимента
            start_experiment_time = int(np.where(self.TF3[start_point:end_point, 0] == 1)[0][0] + start_point) / 250

            # выборка точки времени поднятия руки
            hand_lifting_time = int(np.where(self.TF3[start_point:end_point, 2] == 1)[0][0] + start_point) / 250

            # выборка точки времени поднятия объекта
            object_lifting_indices = np.where(self.TF3[start_point:end_point, 4] == 1)[0]
            if object_lifting_indices.size > 0:
                object_lifting_time = int(object_lifting_indices[0] + start_point) / 250
            else:
                object_lifting_time = np.nan

            # выборка момента открытия пальцев - GA Opening
            ga_opening = int(np.where(self.TF3[start_point:end_point, 3] == 1)[0][0] + start_point) / 250

            # выборка момента открытия пальцев - GA max
            ga_max = int(np.where(self.TF3[start_point:end_point, 3] == 1)[0][0] + start_point) / 250

            # выборка момента опускания объекта - Object Placed
            object_placed_indices = np.where(self.TF3[start_point:end_point, 6] == 1)[0]
            if object_placed_indices.size > 0:
                object_placed = int(object_placed_indices[0] + start_point) / 250
            else:
                object_placed = np.nan

            #["№", "Start Experiment (s)", "Hand Lifting (s)", "QA Opening", "Object Lifting (s)", "QA max", "Object placed"]

            # Добавляем данные эксперимента в список experiment_data
            self.experiment_data.append({
                'Experiment No': idx + 1,
                'Start Experiment': start_experiment_time,
                'Hand Lifted Time': hand_lifting_time,
                'QA Opening': ga_opening,
                'Object Lifted Time': object_lifting_time,
                'GA Max':ga_max, 
                'Object Placed': object_placed
            })

    def calculate_parameters(self):
        
        frame = self.data['frame']
        glasses = self.data['glasses']
        index = self.data['index']
        thumb = self.data['thumb']
        inside = self.data['inside']
        outside = self.data['outside']
        object_ = self.data['object']

        nrows = len(frame)
        FG = np.zeros((nrows, 3))
        GA = np.zeros((nrows, 3))
        Wrist = np.zeros((nrows, 3))

        for t in range(nrows):
            FG[t, 0] = np.sqrt((frame[t, 2] - glasses[t, 2]) ** 2 +
                               (frame[t, 3] - glasses[t, 3]) ** 2 +
                               (frame[t, 4] - glasses[t, 4]) ** 2)
            GA[t, 0] = np.sqrt((index[t, 2] - thumb[t, 2]) ** 2 +
                               (index[t, 3] - thumb[t, 3]) ** 2 +
                               (index[t, 4] - thumb[t, 4]) ** 2)

        Wrist[:, 0] = (inside[:, 2] + outside[:, 2]) / 2
        Wrist[:, 1] = (inside[:, 3] + outside[:, 3]) / 2
        Wrist[:, 2] = (inside[:, 4] + outside[:, 4]) / 2

        self.FG = np.zeros((nrows, 3))
        self.GA = np.zeros((nrows, 3))
        self.Wrist = np.zeros((nrows, 2))

        self.FG[:, 0] = FG[:, 0]
        self.FG[:, 1] = gaussian_filter1d(FG[:, 0], sigma=10)
        self.GA[:, 0] = GA[:, 0]
        self.GA[:, 1] = gaussian_filter1d(GA[:, 0], sigma=10)
        self.Wrist[:, 0] = gaussian_filter1d(Wrist[:, 1], sigma=10)
        self.Inside_Y_Smooth = gaussian_filter1d(inside[:, 3], sigma=10)
        self.Obj_smooth = gaussian_filter1d(object_[:, 3], sigma=30)

        self.FG_speed = np.diff(self.FG[:, 1]) * 100
        self.FG_speed = np.append(self.FG_speed, 0)
        self.Wr_speed = np.diff(self.Wrist[:, 0]) * 100
        self.Wr_speed = np.append(self.Wr_speed, 0)
        self.Obj_speed = np.diff(self.Obj_smooth) * 100
        self.Obj_speed = np.append(self.Obj_speed, 0)

        self.FG_speed = gaussian_filter1d(self.FG_speed, sigma=10)
        self.Obj_speed = gaussian_filter1d(self.Obj_speed, sigma=10)

        self.frame = frame
        self.TF3 = np.zeros((nrows, 8))

    # Разделение массива на эксперименты
    def create_experiment_array(self):
        self.experiment_array = []

        # Находим индексы всех строк, где в первом столбце массива TF3 стоит 1 (начало эксперимента)
        start_indices = np.where(self.TF3[:, 0] == 1)[0]
        # Проходим по каждому индексу начала эксперимента
        for start_point in start_indices:
            adjusted_start = max(0, start_point - 200)  # Уменьшаем начало на 200 кадров, не выходя за пределы массива
            end_point = start_point + 4000  # Конечная точка эксперимента (начало + 4000 кадров)
            # Добавляем кортеж (начало, конец) в массив экспериментов
            self.experiment_array.append((adjusted_start, end_point))
        
    def moment_1_start_experiment(self):

        tl = self.tl # Порогое значение расстояния между пиками
        # Параметры оганичения для FG
        setting_max_FG = self.max_FG 
        setting_min_FG = self.min_FG  

        # Предварительное нахождение пиков
        all_peaks, _ = find_peaks(self.FG_speed, prominence=0.02, distance=tl)
        # Список для сохранения валидных пиков
        valid_peaks = []

        # Проходим по найденным пикам
        for peak in all_peaks:
            fg_value = self.FG[peak, 1]
            # Проверка значения FG на соответствие ограничениям
            if setting_min_FG <= fg_value <= setting_max_FG:
                valid_peaks.append(peak)  # Если пик валиден, добавляем его в список
                print(f"Пик принят, индекс: {peak} FG = {fg_value}")
            else:
                print(f"Пик отклонен, индекс: {peak} FG = {fg_value}")

        # Если не найдено ни одного валидного пика
        if not valid_peaks:
            print("Не найдено ни одного подходящего пика.")
            return

        # Обработка валидных пиков
        for peak in valid_peaks:
            temp2 = peak
            # Поиск первого нулевого значения скорости после максимума
            while temp2 < len(self.FG_speed) and self.FG_speed[temp2] > 0:
                temp2 += 1

            # Записываем найденную точку
            if temp2 < len(self.FG_speed):  # Проверка, чтобы не выйти за пределы массива
                self.TF3[temp2, 0] = 1
                #print(f"Точка начала эксперимента найдена на индексе: {temp2}")

    def moment_2_hand_lifting(self):
        tl = self.hand_lifting # пороговое значение скорости руки
        
        k2 = np.where(self.TF3[:, 0] == 1)[0]
        for temp in k2:
            temp2 = temp
            while True:
                temp2 += 1
                Wr_speed_next = np.mean(self.Wr_speed[temp2:temp2 + 20])
                if self.Wr_speed[temp2] > 0 and Wr_speed_next > tl:
                    break
            self.TF3[temp2, 2] = 1

    def moment_3_finger_opening(self):
        tl = self.finger_opening # Пороговое Значение для открытия пальца
        
        range_ = 100
        k2 = np.where(self.TF3[:, 0] == 1)[0]
        for temp in k2:
            temp2 = temp
            GA_Av = np.mean(self.GA[temp2 - range_:temp2, 1])
            while self.GA[temp2, 1] <= GA_Av * tl:
                temp2 += 1
            self.TF3[temp2, 3] = 1

    def moment_4_object_lifting(self):
        tl = self.object_lifting # Пороговое значение скорости подъема обьекта
        
        k2 = np.where(self.TF3[:, 0] == 1)[0]
        for temp in k2:
            temp2 = temp
            while True:
                temp2 += 1
                # Проверка, чтобы temp2 не выходил за пределы массива
                if temp2 >= len(self.Obj_speed) - 20:  # Убедитесь, что есть достаточно элементов для среднего
                    print("Индекс temp2 выходит за пределы массива")
                    return
                Obj_speed_next = np.mean(self.Obj_speed[temp2:temp2 + 20])
                if self.Obj_speed[temp2] > 0 and Obj_speed_next > tl:
                    break
            self.TF3[temp2, 4] = 1

    def moment_5_max_ga(self):
        # Найти все пики апертуры захвата
        GAmax, _ = find_peaks(self.GA[:, 1], prominence=0.002, distance=100)

        # Найти моменты начала (TF3[:, 2] == 1) и конца (TF3[:, 4] == 1)
        k = np.where(self.TF3[:, 2] == 1)[0]
        k3 = np.where(self.TF3[:, 4] == 1)[0]

        for i, temp in enumerate(k):
            # Убедиться, что момент подъема объекта существует
            lift_idx = k3[i] if i < len(k3) else None

            # Получаем все индексы точек раскрытия апертуры для текущего момента
            open_finger_idx = np.where(self.TF3[:, 3] == 1)[0]  # Все позиции точек раскрытия апертуры

            # Инициализируем список пиков в пределах 800 кадров от точки раскрытия апертуры
            peaks_in_range = []

            for x in GAmax:
                if x >= temp:  # Убедиться, что пик находится после начала движения
                    # Проверяем, что пик находится перед точкой подъема объекта
                    if lift_idx is not None and x < lift_idx:
                        # Проверяем, что хотя бы одна точка раскрытия апертуры находится в пределах 800 кадров
                        if any(abs(x - open_finger) <= 800 for open_finger in open_finger_idx):
                            peaks_in_range.append(x)

            # Если найдены пики, выбираем самый высокий
            if peaks_in_range:
                # Находим индекс максимального пика по значению в GA
                max_peak = max(peaks_in_range, key=lambda x: self.GA[x, 1])
                self.TF3[max_peak, 5] = 1

    def moment_6_object_lowering(self):
        tl = self.object_lowering # Пороговое значение скорости опускания предмета
        
        k2 = np.where(self.TF3[:, 4] == 1)[0]
        for temp in k2:
            temp2 = temp
            while True:
                temp2 += 1
                Obj_speed_next = np.mean(np.abs(self.Obj_speed[temp2:temp2 + 75]))
                if np.abs(self.Obj_speed[temp2]) <= tl / 5 and Obj_speed_next <= tl:
                    break
            self.TF3[temp2, 6] = 1

    def plot_results(self): # Визуализация общего графика
            t = self.frame[:, 1]
            A = self.FG[:, 1]
            B = self.GA[:, 1]
            W = self.Wrist[:, 0]
            Ob = self.Obj_smooth

            Z = self.TF3[:, 0].astype(bool)
            Z2 = self.TF3[:, 2].astype(bool)
            Z3 = self.TF3[:, 3].astype(bool)
            Z4 = self.TF3[:, 4].astype(bool)
            Z5 = self.TF3[:, 5].astype(bool)
            Z6 = self.TF3[:, 6].astype(bool)

            plt.plot(t, A, label='FG')
            plt.plot(t[Z], A[Z], 'b*', label='Start Experiment')
            plt.plot(t, W, label='Wrist_Y')
            plt.plot(t[Z2], W[Z2], 'k*', label='Hand Lifting')
            plt.plot(t, B, label='GA') 
            plt.plot(t[Z5], B[Z5], 'm*', label='Max GA')
            plt.plot(t, Ob, label='Object Y')
            plt.plot(t[Z4], Ob[Z4], 'r*', label='Object Lifting')
            plt.plot(t[Z6], Ob[Z6], 'g*', label='Object Lowering')
            plt.plot(t[Z3], B[Z3], 'y*', label='Finger Opening')

            plt.xlabel('Time (s)')
            plt.ylabel('Distance (cm)')
            plt.legend()
            plt.show()
        
    def plot_results_experiment(self, experiment_index):
        '''Строим график'''
        try:
            # Получаем начальную и конечную точки эксперимента
            start_point, end_point = self.experiment_array[experiment_index]

            # Проверяем, что end_point не выходит за пределы данных
            if end_point > len(self.frame):
                end_point = len(self.frame)

            # Извлекаем данные для данного эксперимента
            t = self.frame[start_point:end_point, 1]  # Время
            A = self.FG[start_point:end_point, 1]  # FG
            B = self.GA[start_point:end_point, 1]  # GA
            W = self.Wrist[start_point:end_point, 0]  # Wrist_Y
            Ob = self.Obj_smooth[start_point:end_point]  # Object Y

            # Извлечение соответствующих точек событий для эксперимента
            Z = self.TF3[start_point:end_point, 0].astype(bool)  # Start Experiment
            Z2 = self.TF3[start_point:end_point, 2].astype(bool)  # Hand Lifting
            Z3 = self.TF3[start_point:end_point, 3].astype(bool)  # Finger Opening
            Z4 = self.TF3[start_point:end_point, 4].astype(bool)  # Object Lifting
            Z5 = self.TF3[start_point:end_point, 5].astype(bool)  # Max GA
            Z6 = self.TF3[start_point:end_point, 6].astype(bool)  # Object Lowering

            # Очищаем фигуру перед построением нового графика
            self.figure.clear()
            ax = self.figure.add_subplot(111)

            # Построение графиков для данного эксперимента
            if self.checkBox_9.isChecked():
                ax.plot(t, A, label='Открытие очков') # FG
            if self.checkBox.isChecked(): # Start Experiment
                ax.plot(t[Z], A[Z], 'b*', label='Начало эксперимента')
            if self.checkBox_4.isChecked(): # Deviation 
                ax.plot(t, W, label='Положение кисти (Z)') # Wrist_Y
            if self.checkBox_2.isChecked(): # Hand Lifting
                ax.plot(t[Z2], W[Z2], 'k*', label='Подъем кисти')
            if self.checkBox_10.isChecked():
                ax.plot(t, B, label='Апертура захвата') # GA 
            if self.checkBox_3.isChecked(): # Max Aperture (Max GA)
                ax.plot(t[Z5], B[Z5], 'm*', label='Апертура захвата (Max)')
            if self.checkBox_8.isChecked(): # Trajectory
                ax.plot(t, Ob, label='Положение объекта (Z)')
            if self.checkBox_5.isChecked(): # Object Lifting
                ax.plot(t[Z4], Ob[Z4], 'r*', label='Подъем объекта')
            if self.checkBox_6.isChecked(): # Object Lowering
                ax.plot(t[Z6], Ob[Z6], 'g*', label='Опускание объекта')
            if self.checkBox_7.isChecked(): # Finger Opening
                ax.plot(t[Z3], B[Z3], 'y*', label='Размыкание пальцев')
            ax.set_xlabel('Time (s)')
            ax.set_ylabel('Distance (cm)')
            ax.legend()
            ax.set_title(f'Experiment {experiment_index + 1}')

            # Обновляем график
            self.canvas.draw()
        except:
            QMessageBox.warning(self, "Ошибка", "Отсутствуют данные для обновления графика")
            return

    def edit_excel(self, experiment_index):
        '''Редактирование Excel файла'''
        # Загрузка данных из Excel 
        excel_df = pd.read_excel(self.file_path2)

        # Копия DataFrame
        modified_df = excel_df.copy()

        # Переименовывем столбцы
        modified_df = modified_df.rename(columns={'Unnamed: 0': 'Subject', 'Figure':'Object', 'Figure \nOrientation':'Object orientation', 'Plate orientation':'Plate orientation'})  

        # Удаляем ненужные столбцы
        modified_df = modified_df.drop(columns=['Unnamed: 4', 'Unnamed: 5', 'Unnamed: 6'])

        # Добавление новых столбец
        modified_df['Object Degree'] = (modified_df['Object orientation']-1)*90
        modified_df['Plate Degree'] = (modified_df['Plate orientation'] - 1)*90
        modified_df['Rotation'] = modified_df['Object Degree'] - modified_df['Plate Degree']

        # Формирование столбца Group
        def calculate_group(rotation):
            if rotation == 0:
                return "N1"
            elif rotation == -90:
                return "N4"
            elif rotation == 90:
                return "N2"
            elif rotation == 180:
                return "N3"
            elif rotation == -180:
                return "N3"
            elif rotation == 270:
                return "N4"
            elif rotation == -270:
                return "N2"

        modified_df['Group'] = modified_df['Rotation'].apply(calculate_group)

        modified_df['Subject'] = experiment_index # Изменяем индексы на номер эксперимента

        print(modified_df.columns)
        return modified_df

    def export_data_to_excel(self):
        """Экспортирует данные эксперимента в Excel с выбором пути сохранения."""
        if self.file_loaded == False:
            QMessageBox.warning(self, "Ошибка", "Отсутсвует файл с данными испытания")
            return

        experiment_index, ok = QInputDialog.getText(self, "Введите номер испытания", "1. Неоходимо ввести единый индекс для экспериментов данного испытания и нажать 'ОК' ниже. (Пример названия: S1) \n2. Далее выбрать файл шаблона с информацией по испытанию и нажать 'Открыть' в диалоговом окне проводника. \n3. В следующем открывшемся окне задать путь сохранения файла и нажать 'Сохранить' в диалоговом окне проводника. \nНомер испытания: ")
        if not ok or not experiment_index:
            QMessageBox.warning(self, "Ошибка", "Номер испытания не был введен")
            return
        
        self.file_path2, _ = QFileDialog.getOpenFileName(self, "Exsel files", "*.xlsx")
        if not self.file_path2:
            QMessageBox.warning(self, "Ошибка", "Не выбран файл с информацией об испытании")
            return

        # Открытие диалога для выбора места сохранения файла
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                    filetypes=[("Excel files", "*.xlsx"),
                                                            ("All files", "*.*")])
        if not file_path:
            return  # Если путь не выбран, выходим из функции

        try:
            # Создание DataFrame из данных self.experiment_data
            mat_df = pd.DataFrame(self.experiment_data,
                                columns=['Experiment No', 'Start Experiment', 'Hand Lifted Time', 'QA Opening', 'Object Lifted Time', 'GA Max', 'Object Placed'])
            
            mat_df = mat_df.rename(columns={'Start Experiment': 'Start', 'Hand Lifted Time':'Hand lifting', 'Object Lifted Time':'Object lifting', 'QA Opening':'GA opening'})
            # Загрузка данных из редактированного Excel файла
            modified_df = self.edit_excel(experiment_index)

            # Объединение mat и Excel данных 
            combined_df = pd.concat([modified_df, mat_df.iloc[:, 1:]], axis=1)

            total_df = self.edit_total_for_export(combined_df)
            
            # Запись DataFrame в файл Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                total_df.to_excel(writer, index=False)

            workbook = load_workbook(file_path)
            worksheet = workbook.active

            header_font = Font(bold=True) 
            header_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid') 
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) 

            # Проходимся по первой строке (заголовкам)
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            # Увеличиваем высоту первой строки
            worksheet.row_dimensions[1].height = 30

            # Добавляем тонкие границы для всех ячеек
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

            # Добавляем жирные линии на строках 6, 8 и 10
            thick_border = Border(right=Side(style='thick'))
            for column_number in [8, 14, 24]:
                for row in worksheet.iter_rows():
                    cell = row[column_number - 1]
                    cell.border = Border(left=cell.border.left, right=Side(style='thick'), top = cell.border.top, bottom= cell.border.bottom)

            # Сохраняем файл с применёнными стилями
            workbook.save(file_path) 

            # Уведомление об успешном экспорте
            QMessageBox.information(self, "Экспорт завершен", f"Таблица успешно сохранена в {file_path}")
        except Exception as e:
            # Вывод сообщения об ошибке, если что-то пошло не так
            QMessageBox.warning(self, "Ошибка экспорта", f"Не удалось сохранить файл. (Необходимо выбрать корректный файл )")

    def edit_total_for_export(self, combined_df):
        '''Собираем страничку Total для экспортируемого файла'''

        combined_df['Hand\n lifting'] = combined_df['Hand lifting'] - combined_df['Start']
        combined_df['GA\n opening'] = combined_df['GA opening'] - combined_df['Start']
        combined_df['Object\n lifting'] = combined_df['Object lifting'] - combined_df['Start']
        combined_df['GA max\n'] = combined_df['GA Max'] - combined_df['Start']
        combined_df['Total movement time'] = combined_df['Object Placed'] - combined_df['Start']
        combined_df['Reaction time'] = combined_df['Hand\n lifting']
        combined_df['Time of max GA'] = combined_df['GA max\n'] - combined_df['Hand\n lifting']
        combined_df['Time to reach'] = combined_df['Object\n lifting'] - combined_df['Hand\n lifting']
        combined_df['GA%'] = (combined_df['Time of max GA']/combined_df['Time to reach']) * 100
        combined_df['Time of object movement'] = combined_df['Total movement time'] - combined_df['Object\n lifting']

        return combined_df


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())